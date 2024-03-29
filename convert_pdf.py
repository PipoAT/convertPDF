import PySimpleGUI as sg
import pdfplumber
from openpyxl import load_workbook
import pandas as pd
import os

def excel_to_prgm(wbactivesheet):
    """
    excel_to_prgm reads the data from the created excel spreadsheet and creates the define
    attributes in a newly created c plus plus file.

    :param wbactivesheet: the active excel sheet that contains the data
    """
    # open a text file in write mode
    with open('output.cpp', 'w') as file:
        # get the headers from the first row of the worksheet
        headers = [cell.value for cell in wbactivesheet[1]]

        # check if 'address' and 'page' are in the headers and get their indices
        indices_to_exclude = [i for i, header in enumerate(headers) if 'address' in header.lower() or 'page' in header.lower()]

        # iterate through each row in the worksheet starting from the second row
        for row in wbactivesheet.iter_rows(min_row=2, values_only=True):
            # skip rows without a name
            if row[0] is None: continue
            # Write #define directives to cpp file
            for i, cell in enumerate(row[2:10], 1):
                # skip the cells with invalid data
                if i in indices_to_exclude or not cell or cell in ['-', '—', ' ']:
                    continue
                cell = cell.split('(')[0] if '(' in cell else cell
                # prevent large amount of text from showing
                if len(cell) > 15: continue
                 # checks if the cell contains PIN, PORT, or DD to specifically format the resulting cpp define directive
                if 'PIN' in cell:
                    file.write(f'#define {cell} {cell[3]}, {cell[4]}\n')
                elif 'PORT' in cell:
                    file.write(f'#define {cell} {cell[4]}, {cell[5]}\n')
                elif 'DD' in cell:
                    file.write(f'#define {cell} {cell[2]}, {cell[3]}\n')
                else: # if none of the above is true, define with original data/info
                    file.write(f'#define {row[1]}_Bit{8-i} {cell}\n')

    
def merge_cells_with_text(excel_file, sheet_name):
    '''
    merge_cells_with_text is a function that merges cells in an Excel file
    which include the one with any non-empty text followed by any and all cells
    that are blank to the right of that cell

    :param excel_file: the excel file that contains data/cells that need formatted
    :param sheet_name: the active sheet within the excel file that contains data/cells that need formatted
    '''
    # Load the workbook and select the sheet
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    # Iterate over the rows in the sheet
    for row in ws.iter_rows():
        # Iterate over the cells in the row
        for cell in row:
            # Get the value of the current cell as a stripped string (empty string if None)
            value = str(cell.value).strip() if cell.value is not None else ''
            # Check if the current cell contains non-empty text
            if value:
                # Find the next non-empty cell in the row
                next_cell_index = cell.column + 1
                while next_cell_index <= ws.max_column:
                    # Get the value of the next cell as a stripped string (empty string if None)
                    next_cell = ws.cell(row=cell.row, column=next_cell_index).value
                    # If the next cell contains non-empty text, exit the loop
                    if next_cell is not None:
                        break
                    next_cell_index += 1
                # Merge the cells from the current cell to the next non-empty cell
                ws.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row, end_column=next_cell_index - 1)

    # Save the workbook
    wb.save(excel_file)
    # convert to .cpp file with define statements
    excel_to_prgm(wb.active)


def pdf_to_excel(pdf_file, page_numbers):
    '''
    pdf_to_excel converts the pdf tables into the excel spreadsheet with desired formatting of
    merged cells and certain columns

    :param pdf_file: the pdf file that the python script will read from to obtain data
    :param page_numbers: user-defined pages of a pdf as an integer(s) that the script will locate data from within the pdf
    '''
    # obtain file names, folder, and pages
    pdf_file_name = os.path.splitext(os.path.basename(pdf_file))[0]
    folder_selected = os.path.dirname(pdf_file)
    excel_file = os.path.join(folder_selected, pdf_file_name[:10] + ".xlsx")

    has_tables = False
    with pdfplumber.open(pdf_file) as pdf:
        try: # check if the pages are within the range of the pdfr
            page = pdf.pages[page_numbers-1]
        except IndexError: # handle index/out of range error by setting page to 0 to indicate no tables exist
            page = pdf.pages[0]
        tables = page.extract_tables()
        # checks if there is existing data
        if any(cell.strip() for table in tables for row in table for cell in row):
            has_tables = True
            # iterates through each table and formats the data in the excel spreadsheet
            for j, table in enumerate(tables):
                df = pd.DataFrame(table[1:], columns=table[0])

                try:
                    # Exclude rows with reserved names or '-'
                    df = df[(df[df.columns[1]] != 'Reserved') & (df[df.columns[1]] != '—')]
                except IndexError: # if there are any errors with exclusion, ignore and continue on
                    continue

                # Convert numeric values to numeric type
                df = df.apply(pd.to_numeric, errors='ignore')
                sheet_name = f"{pdf_file_name[:10]}_Page{page_numbers}_Table{j+1}"
                with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        else: # If no tables exist, throw alert.
            sg.popup("ALERT", "No Tables Exist.")

    merge_cells_with_text(excel_file, f"{pdf_file_name[:10]}_Page{page_numbers}_Table{j+1}") if has_tables else ""

# define the layout of the GUI
layout = [
    [sg.Text("Select PDF File:")],
    [sg.Input('', key="file"), sg.FileBrowse(file_types=(("PDF Files", "*.pdf"),))],
    [sg.Text("Enter the page number: ")],
    [sg.Input('', key="page")],
    [sg.Button("Convert"), sg.Button("Exit"), sg.Button("About")]
]

# create the window
window = sg.Window("PDF Conversion GUI", layout)

while True:
    # display and interact with window
    event, values = window.read()

    # if user closes window or clicks exit
    if event == sg.WINDOW_CLOSED or event == 'Exit':
        break

    # perform action with information
    if event == "Convert":
        # obtain the pdf file from user input
        pdf_file = values['file']
        # obtain page numbers from user input
        page_numbers = int(values['page']) if values['page'].isdigit() else 0
        # check if the page numbers and pdf file are valid inputs by user
        if not pdf_file:
            sg.popup("ALERT","Please select a PDF file")
        elif page_numbers <= 0:
            sg.popup("ALERT","Please enter valid page numbers!")
        else:
            # call the function to convert
            pdf_to_excel(pdf_file, page_numbers)
            # show alert that conversion is complete and clear input fields
            sg.popup("ALERT","Process Completed!")
            window['page'].update('')
    
    # if user selects the About button, display the software info as a popup/alert
    sg.popup("INFO","Version 1.0.0\nSoftware Developed by Andrew T. Pipo") if event == "About" else ""

# close window
window.close()